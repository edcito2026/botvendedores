#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WEBHOOK RESPUESTA A MENSAJE "TROYA" - V1

Escucha mensajes WhatsApp entrantes.
Si detecta palabra clave "TROYA", envía reporte personalizado de clientes TROYA.

Flujo:
1. Vendedor envía: "TROYA"
2. Webhook recibe el mensaje
3. Identifica teléfono del vendedor
4. Obtiene clientes TROYA (Calif='D') del vendedor
5. Genera mensaje personalizado
6. Envía respuesta automática

Variables de entorno:
WHATSAPP_ACCESS_TOKEN
WHATSAPP_PHONE_NUMBER_ID
WHATSAPP_VERIFY_TOKEN
WHATSAPP_API_VERSION (default v25.0)
BD_PATH (default ventas.db)
EXCEL_VENDEDORES (default vendedores.xlsx)
"""

from flask import Flask, request, jsonify
import sqlite3
import requests
import logging
import os
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
from threading import Lock
import time

# ============================================================
# 1. CONFIGURACIÓN
# ============================================================

BASE_PATH = os.path.dirname(os.path.abspath(__file__))

ACCESS_TOKEN = os.environ.get("WHATSAPP_ACCESS_TOKEN")
PHONE_NUMBER_ID = os.environ.get("WHATSAPP_PHONE_NUMBER_ID")
VERIFY_TOKEN = os.environ.get("WHATSAPP_VERIFY_TOKEN")
API_VERSION = os.environ.get("WHATSAPP_API_VERSION", "v25.0")

if not ACCESS_TOKEN:
    raise RuntimeError("Falta WHATSAPP_ACCESS_TOKEN")
if not PHONE_NUMBER_ID:
    raise RuntimeError("Falta WHATSAPP_PHONE_NUMBER_ID")
if not VERIFY_TOKEN:
    raise RuntimeError("Falta WHATSAPP_VERIFY_TOKEN")

API_URL = f"https://graph.facebook.com/{API_VERSION}/{PHONE_NUMBER_ID}/messages"

BD_PATH = os.environ.get("BD_PATH", os.path.join(BASE_PATH, "ventas.db"))
EXCEL_VENDEDORES = os.environ.get(
    "EXCEL_VENDEDORES",
    os.path.join(BASE_PATH, "vendedores.xlsx")
)

# Zona horaria Perú
try:
    TZ_LIMA = ZoneInfo("America/Lima")
except Exception:
    TZ_LIMA = None

# ============================================================
# 2. LOGGING
# ============================================================

LOG_DIR = os.path.join(BASE_PATH, "logs")
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(
            os.path.join(LOG_DIR, "webhook_troya_respuesta.log"),
            encoding="utf-8"
        ),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Cache Excel con Lock
_excel_cache = {}
_excel_cache_lock = Lock()
_excel_mtime = 0


# ============================================================
# 3. UTILIDADES
# ============================================================

def ahora_local():
    """Fecha/hora actual en Perú."""
    if TZ_LIMA:
        return datetime.now(TZ_LIMA)
    return datetime.now()


def normalizar_telefono(numero):
    """Normaliza teléfono a formato 51XXXXXXXXX"""
    if not numero:
        return None

    # Eliminar caracteres no numéricos
    limpio = ''.join(filter(str.isdigit, numero))

    # Si ya comienza con 51, retornar
    if limpio.startswith('51'):
        return limpio

    # Si comienza con 0, reemplazar con 51
    if limpio.startswith('0'):
        return f"51{limpio[1:]}"

    # Si tiene 9 dígitos (caso Perú sin prefijo), agregar 51
    if len(limpio) == 9:
        return f"51{limpio}"

    return limpio


def obtener_ultimo_numero(numero):
    """Obtiene los últimos 9 dígitos del número"""
    limpio = ''.join(filter(str.isdigit, numero))
    if len(limpio) >= 9:
        return limpio[-9:]
    return limpio


def cargar_vendedores_excel():
    """Carga vendedores del Excel con caché"""
    global _excel_cache, _excel_mtime

    try:
        if not os.path.exists(EXCEL_VENDEDORES):
            logger.error(f"❌ Excel no encontrado: {EXCEL_VENDEDORES}")
            return {}

        # Verificar si el archivo cambió
        mtime = os.path.getmtime(EXCEL_VENDEDORES)

        with _excel_cache_lock:
            if _excel_cache and mtime == _excel_mtime:
                return _excel_cache

            # Recargar
            vendedores = {}
            wb = openpyxl.load_workbook(EXCEL_VENDEDORES)
            ws = wb['Vendedores'] if 'Vendedores' in wb.sheetnames else wb.active

            for row in ws.iter_rows(min_row=4, values_only=True):
                if row[0] is None:
                    break

                codigo = row[0]
                nombre = row[1]
                telefono = row[2]

                if codigo and nombre and telefono:
                    tel_norm = normalizar_telefono(str(telefono))
                    tel_9dig = obtener_ultimo_numero(str(telefono))

                    vendedores[tel_norm] = {
                        'codigo': codigo,
                        'nombre': str(nombre).strip(),
                        'telefono': tel_norm
                    }

                    # Agregar también por últimos 9 dígitos
                    vendedores[tel_9dig] = vendedores[tel_norm]

            _excel_cache = vendedores
            _excel_mtime = mtime
            logger.info(f"✅ {len(set(v['codigo'] for v in vendedores.values()))} vendedores cargados")
            return vendedores

    except Exception as e:
        logger.error(f"❌ Error leyendo Excel: {e}")
        return {}


def obtener_clientes_troya(cod_vendedor):
    """Obtiene clientes TROYA (Calif='D') del vendedor"""
    try:
        conn = sqlite3.connect(BD_PATH, timeout=15)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Obtener clientes CALIF=D
        cursor.execute("""
            SELECT
                Cod_Clie,
                Raz_Social,
                Giro
            FROM clientes
            WHERE Calif = 'D' AND Cdg_Vend = ?
            ORDER BY Raz_Social
        """, (cod_vendedor,))

        clientes = [dict(row) for row in cursor.fetchall()]

        # Obtener período actual para consulta de compras
        ahora = ahora_local()
        periodo = f"{ahora.year}{ahora.month:02d}"
        tabla_ventas = f"VENTAS{ahora.year}"

        # Verificar compras para cada cliente
        for cliente in clientes:
            try:
                cursor.execute(f"""
                    SELECT COUNT(*) as total_compras
                    FROM {tabla_ventas}
                    WHERE Cod_Clie = ? AND strftime('%Y%m', Fecha) = ?
                """, (cliente['Cod_Clie'], periodo))

                resultado = cursor.fetchone()
                total_compras = resultado['total_compras'] if resultado else 0
                cliente['tiene_compras'] = total_compras > 0
            except:
                cliente['tiene_compras'] = False

        conn.close()
        return clientes

    except Exception as e:
        logger.error(f"❌ Error consultando clientes TROYA: {e}")
        return []


def generar_mensaje_troya(vendedor, clientes):
    """Genera mensaje de respuesta para solicitud TROYA"""

    if not clientes:
        return f"""Hola {vendedor['nombre']}, 👋

No tienes clientes TROYA (Calif=D) registrados actualmente.

🤖 Bot N&J"""

    nombre = vendedor['nombre'].strip().title()
    cantidad = len(clientes)
    con_compra = sum(1 for c in clientes if c['tiene_compras'])
    sin_compra = cantidad - con_compra

    ahora = ahora_local()
    mes_nombre = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                  "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"][ahora.month - 1]

    mensaje = f"""👤 HOLA {nombre}
📅 {mes_nombre} {ahora.year}

📊 TUS CLIENTES TROYA:
Total: {cantidad} clientes especiales
✅ {con_compra} CON COMPRA (activos)
❌ {sin_compra} SIN COMPRA (reactivar)

"""

    # Clientes CON COMPRA
    con_compra_list = [c for c in clientes if c['tiene_compras']]
    if con_compra_list:
        mensaje += "✅ CON COMPRA (MANTENER):\n"
        for idx, cliente in enumerate(con_compra_list, 1):
            cliente_nombre = cliente['Raz_Social'].strip().title()
            giro = cliente['Giro'].strip().lower() if cliente['Giro'] else ''
            mensaje += f"{idx}. {cliente_nombre} ({giro})\n"

    # Clientes SIN COMPRA
    sin_compra_list = [c for c in clientes if not c['tiene_compras']]
    if sin_compra_list:
        mensaje += "\n❌ SIN COMPRA (REACTIVAR):\n"
        for idx, cliente in enumerate(sin_compra_list, 1):
            cliente_nombre = cliente['Raz_Social'].strip().title()
            giro = cliente['Giro'].strip().lower() if cliente['Giro'] else ''
            mensaje += f"{idx}. {cliente_nombre} ({giro})\n"

    mensaje += """
📈 ACCIONES:
1️⃣ Contacta activos
2️⃣ Reactiva sin compra
3️⃣ Envía promociones
4️⃣ Cierra la venta

🤖 Bot N&J"""

    return mensaje


def enviar_respuesta_api(numero_destino, mensaje_texto):
    """Envía respuesta por WhatsApp Business API"""
    try:
        payload = {
            "messaging_product": "whatsapp",
            "to": numero_destino,
            "type": "text",
            "text": {
                "body": mensaje_texto
            }
        }

        headers = {
            "Authorization": f"Bearer {ACCESS_TOKEN}",
            "Content-Type": "application/json"
        }

        response = requests.post(API_URL, json=payload, headers=headers, timeout=10)

        if response.status_code == 200:
            logger.info(f"✅ Respuesta enviada a {numero_destino}")
            return True
        else:
            logger.error(f"❌ Error API ({response.status_code}): {response.text}")
            return False

    except Exception as e:
        logger.error(f"❌ Error enviando: {e}")
        return False


# ============================================================
# 4. WEBHOOK ROUTES
# ============================================================

@app.route("/webhook", methods=["GET"])
def webhook_verify():
    """Verifica el webhook con Meta"""
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")

    if mode == "subscribe" and token == VERIFY_TOKEN:
        logger.info("✅ Webhook verificado con Meta")
        return challenge, 200
    else:
        logger.warning("❌ Verificación de webhook fallida")
        return jsonify({"status": "error"}), 403


@app.route("/webhook", methods=["POST"])
def webhook_receive():
    """Recibe mensajes WhatsApp entrantes"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({"status": "ok"}), 200

        # Obtener mensaje
        messages = data.get("entry", [{}])[0].get("changes", [{}])[0].get("value", {}).get("messages", [])

        if not messages:
            return jsonify({"status": "ok"}), 200

        mensaje_obj = messages[0]
        numero_origen = mensaje_obj.get("from")
        texto_mensaje = mensaje_obj.get("text", {}).get("body", "").strip().upper()

        logger.info(f"📩 Mensaje recibido de {numero_origen}: {texto_mensaje[:50]}")

        # Detectar palabra clave TROYA
        if "TROYA" not in texto_mensaje:
            logger.info(f"⊘ No contiene 'TROYA', ignorado")
            return jsonify({"status": "ok"}), 200

        # Cargar vendedores
        vendedores = cargar_vendedores_excel()

        if not vendedores:
            logger.error("❌ No hay vendedores cargados")
            return jsonify({"status": "error"}), 500

        # Buscar vendedor por teléfono
        vendedor = vendedores.get(numero_origen) or vendedores.get(obtener_ultimo_numero(numero_origen))

        if not vendedor:
            logger.warning(f"⚠️ Vendedor no encontrado: {numero_origen}")
            respuesta = "Hola 👋\n\nNo estás registrado como vendedor en nuestra base de datos.\nContacta al equipo de N&J.\n\n🤖 Bot N&J"
            enviar_respuesta_api(numero_origen, respuesta)
            return jsonify({"status": "ok"}), 200

        logger.info(f"👤 Vendedor identificado: {vendedor['nombre']} ({vendedor['codigo']})")

        # Obtener clientes TROYA
        clientes_troya = obtener_clientes_troya(vendedor['codigo'])

        # Generar mensaje
        mensaje_respuesta = generar_mensaje_troya(vendedor, clientes_troya)

        # Enviar respuesta
        logger.info(f"📤 Enviando respuesta a {vendedor['nombre']}")
        if enviar_respuesta_api(numero_origen, mensaje_respuesta):
            logger.info(f"✅ Respuesta exitosa para {vendedor['nombre']}")
        else:
            logger.error(f"❌ Error en respuesta para {vendedor['nombre']}")

        return jsonify({"status": "ok"}), 200

    except Exception as e:
        logger.error(f"❌ Error procesando webhook: {e}")
        return jsonify({"status": "error"}), 500


# ============================================================
# 5. HEALTH CHECK
# ============================================================

@app.route("/health", methods=["GET"])
def health():
    """Health check para Render"""
    return jsonify({"status": "ok", "timestamp": ahora_local().isoformat()}), 200


# ============================================================
# 6. MAIN
# ============================================================

if __name__ == "__main__":
    logger.info("="*70)
    logger.info("WEBHOOK TROYA - RESPUESTA AUTOMÁTICA")
    logger.info("="*70)
    logger.info(f"📊 BD: {BD_PATH}")
    logger.info(f"📋 Excel: {EXCEL_VENDEDORES}")
    logger.info(f"🔑 Palabra clave: TROYA")
    logger.info("="*70)

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
